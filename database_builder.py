""" Helper functions for database building, populating, and file uploading for use in app.py """
import sqlite3
import openpyxl

def create_database():
    """
    Creates a sqlite3 database ('timbers_future.db') and table with the following features:
        date TEXT PRIMARY KEY,
        open REAL,
        high REAL,
        low REAL,
        close REAL,
        adj_close REAL,
        volume INTEGER
    
    input: None
    output: None
    """
    
    # Connect to the database
    conn = sqlite3.connect('timbers_future.db')
    cursor = conn.cursor()

    # Create the table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lumber_futures (
            date TEXT PRIMARY KEY,
            open REAL,
            high REAL,
            low REAL,
            close REAL,
            adj_close REAL,
            volume INTEGER) 
        ''')

    # Commit the change
    conn.commit()
    # Close the connection
    conn.close()
    # return conn, cursor

def import_data_to_database():
    """
    Read data from xlsx input file and load to sqlite3 database.

    input: None
    output: None (data loaded to database)

    Designed to run after create_database()
    """
    # Connect to the database
    conn = sqlite3.connect('timbers_future.db')
    cursor = conn.cursor()

    # Open the XLSX file
    workbook = openpyxl.load_workbook('LumberFut.xlsx')
    workbook = workbook[workbook.sheetnames[0]]
    
    # Iterate over the rows of the sheet
    for index, row in enumerate(workbook.iter_rows(min_row=2)):
            # Get the values from the cells
            date = row[0].value
            open = row[1].value
            high = row[2].value
            low = row[3].value
            close = row[4].value
            adj_close = row[5].value
            volume = row[6].value
            
            # Insert the data into the table
            cursor.execute('''
                INSERT OR REPLACE INTO lumber_futures (date, open, high, low, close, adj_close, volume)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (date, open, high, low, close, adj_close, volume))

    # Commit the changes
    conn.commit()
    
    # Close the connection
    conn.close()

def key_values():
    """
    To get key metric values (mix/max) for each value in database
    input : None
    output : [max_open, min_open, max_high, min_high, max_low, min_low, max_close, min_close,
              max_adj_close, min_adj_close, max_volume, min_volume]
              Max and min values for each of [open, high, low, close, adj_close, volume]
    Designed to run after create_database()
    """
    # Connect to the database
    conn = sqlite3.connect('timbers_future.db')
    cursor = conn.cursor()

    # get max and min values for open
    max_open = cursor.execute('''SELECT MAX(open) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    max_open = cursor.fetchone()
    min_open = cursor.execute('''SELECT MIN(open) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    min_open = cursor.fetchone()

    # get max and min values for high
    max_high = cursor.execute('''SELECT MAX(high) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    max_high = cursor.fetchone()
    min_high = cursor.execute('''SELECT MIN(high) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    min_high = cursor.fetchone()

    # get max and min values for low
    max_low = cursor.execute('''SELECT MAX(low) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    max_low = cursor.fetchone()
    min_low = cursor.execute('''SELECT MIN(low) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    min_low = cursor.fetchone()

    # get max and min values for close
    max_close = cursor.execute('''SELECT MAX(close) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    max_close = cursor.fetchone()
    min_close = cursor.execute('''SELECT MIN(close) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    min_close = cursor.fetchone()

    # get max and min values for adj_close
    max_adj_close = cursor.execute('''SELECT MAX(adj_close) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    max_adj_close = cursor.fetchone()
    min_adj_close = cursor.execute('''SELECT MIN(adj_close) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    min_adj_close = cursor.fetchone()

    # get max and min values for volume
    max_volume = cursor.execute('''SELECT MAX(volume) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    max_volume = cursor.fetchone()
    min_volume = cursor.execute('''SELECT MIN(volume) FROM lumber_futures
        WHERE COALESCE(date, open, high, low, close, adj_close, volume) IS NOT NULL 
        AND open NOT LIKE '%-%';''')
    min_volume = cursor.fetchone()

    conn.close()
    return [max_open, min_open, max_high, min_high, max_low, min_low, max_close, min_close,
            max_adj_close, min_adj_close, max_volume, min_volume]